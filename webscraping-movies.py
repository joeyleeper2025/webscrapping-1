
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

table_rows = soup.findAll('tr')




#to excel
wb = xl.Workbook()
ws = wb.active

ws.title = "Box Office Report"

header_font = Font(name='Times New Romans',size=24,italic=False,bold=True)
ws['A1'] = 'No.'
ws['A1'].font = header_font
ws.column_dimensions['A'].width = 15

ws['B1'] = 'Title'
ws['B1'].font = header_font
ws.column_dimensions['B'].width = 30

ws['C1'] = 'Release Date'
ws['C1'].font = header_font
ws.column_dimensions['C'].width = 30

ws['D1'] = 'Number of Theaters'
ws['D1'].font = header_font
ws.column_dimensions['D'].width = 40

ws['E1'] = 'Top Gross'
ws['E1'].font = header_font
ws.column_dimensions['E'].width = 30

ws['F1'] = 'Average gross by theater'
ws['F1'].font = header_font
ws.column_dimensions['F'].width = 50

write_sheet = wb['Box Office Report']


for row in table_rows[1:6]:
    td = row.findAll('td')
    number = td[0].text
    movie = td[1].text
    r_date = td[8].text
    total_gross = int(td[5].text.replace('$','').replace(',',''))
    total_t = int(td[6].text.replace(',',''))

    average_gross = total_gross / total_t





    write_sheet.append([number,movie,r_date,total_t,total_gross,average_gross])


for cell in write_sheet['D':'D']:
        cell.number_format = '#,##0'

for cell in write_sheet['E':'E']:
        cell.number_format = u'"$"#,##0'
    
for cell in write_sheet['F':'F']:
        cell.number_format = u'"$"#,##0'
    

wb.save('Box Office.xlsx')



