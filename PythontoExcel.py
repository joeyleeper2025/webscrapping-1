import openpyxl as xl
from openpyxl.styles import Font

#create a new excel doc

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'


wb.create_sheet(index=1,title='Second Sheet')

#write content to a cell

ws['A1'] = 'Invoice'
ws['A1'].font = Font(name='Times New Romans',size=24,italic=False,bold=True)

header_font = Font(name='Times New Romans',size=24,italic=False,bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Allignment'



#ws.merge_cells['A1' : 'B1']

ws['B2']= 450
ws['B3']= 225
ws['B4']= 150

ws['A8']= 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=SUM(B2:B7)'

ws.column_dimensions['A'].width = 25






# Load the 'ProduceReport.xlsx' workbook and access the 'ProduceReport' sheet
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

# Assuming you have already created write_sheet as 'Second Sheet' in your previous code
write_sheet = wb['Second Sheet']

# Iterate over rows in read_ws, extract values, and append them to write_sheet
for row in read_ws.iter_rows(min_row=1, max_row=read_ws.max_row, max_col=read_ws.max_column, values_only=True):
    write_sheet.append(row)
    
max_row = write_sheet.max_row


# Write totals
write_sheet.cell(max_row + 2, 1).value = 'Totals'
write_sheet.cell(max_row + 2, 3).value = f"=SUM(C1:C{max_row})"
write_sheet.cell(max_row + 2, 4).value = f"=SUM(D1:D{max_row})"

# Write averages
write_sheet.cell(max_row + 3, 1).value = 'Averages:'
write_sheet.cell(max_row + 3, 3).value = f"=AVERAGE(C1:C{max_row})"
write_sheet.cell(max_row + 3, 4).value = f"=AVERAGE(D1:D{max_row})"

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet['C':'C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D':'D']:
    cell.number_format = u'"$"#,##0'




















wb.save('PythontoExcel.xlsx')