import openpyxl as xl


wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

sheet1 = wb['Sheet1']

print(sheet1)

cellA1 = sheet1['A1']

print(cellA1.value)
print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)

print(sheet1.max_row)
print(sheet1.max_column)

# i = 1 
# for row in sheet1:
#     print(sheet1.cell(i , 2).value)

#     i += 1 

# for row in sheet1:
#     fruit = row[1].value
#     print(fruit)


for i in range(1,sheet1.max_row+1):
    print(sheet1.cell(i,2).value)


print(xl.utils.get_column_letter(2))
print(xl.utils.column_index_from_string('A'))

# for current in sheet1['A1':'C3']:
#     for currentcell in current:
#         print(currentcell.coordinate)
#         print(currentcell.value)
#         input()

for current_row in sheet1.iter_rows(min_row=1,max_row=sheet1.max_row,max_col=sheet1.max_column):
    
    print(current_row[0].value)
    print(current_row[1].value)
    print(current_row[2].value)
    print()
    input()
    